#!/usr/bin/python2.7

from openpyxl import load_workbook
import pymongo
from connection import mongo_uri
from location_import import doImport

def p(x):
    try:
        return str(int(x))
    except:
        return None 
def nodeImport():
    doImport()
    with pymongo.MongoClient(mongo_uri) as client:
        db = client.mack0242
        locations = db['locations']
       
        path = "C:\\Users\\mack0242\\Dropbox\\PhD\\scats_data\\ACC Turning Movements.xlsx"
    
        wb = load_workbook(filename=path,use_iterators=True)
        ws = wb.get_sheet_by_name(name='Coding')
        r = 0
        for row in ws.iter_rows():
            #check if the first cell starts with TS
            ts = row[0].internal_value
            if ts and ts.strip()[:2] == "TS":
                
                fromNode = p(row[3].internal_value)
                toNode = p(row[5].internal_value)
                turn = {
                        'type':row[1].internal_value,
                        'query':row[6].internal_value,
                        'movement':row[2].internal_value
                    }
                if fromNode:
                    turn['from'] = fromNode
                if toNode:
                    turn['to']   = toNode
                node =p(row[4].internal_value)
                result = locations.update(
                                 {'intersection_number':node},
                                 {'$push':{'turns':turn}},True)
                if result and result['n']:
                    r+=1
                
        print("Rows updated:",r)
        
if __name__=="__main__":
    nodeImport()